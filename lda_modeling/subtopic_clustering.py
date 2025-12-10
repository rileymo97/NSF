"""
Subtopic clustering functions using semantic embeddings
"""
import os
import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from pathlib import Path
from openpyxl import load_workbook
from scipy.spatial.distance import pdist
from scipy.cluster.hierarchy import linkage, dendrogram, fcluster, inconsistent
from sentence_transformers import SentenceTransformer


def topic_centroid(keyphrases, model):
    """
    Calculate centroid embedding for a list of keyphrases.
    
    Args:
        keyphrases: List of keyphrase strings
        model: SentenceTransformer model
        
    Returns:
        Normalized centroid embedding vector
    """
    keyphrases_clean = [k.replace("_", " ") for k in keyphrases]
    embs = model.encode(keyphrases_clean, normalize_embeddings=True)
    centroid = embs.mean(axis=0)
    
    centroid = centroid / (np.linalg.norm(centroid) + 1e-9)
    return centroid


def plot_similarity_matrix(topic_vectors, topic):
    """
    Plot cosine similarity matrix for topic vectors.
    
    Args:
        topic_vectors: Array of topic embedding vectors
        topic: Division name for plot title
    """
    topic_sim_matrix = topic_vectors @ topic_vectors.T

    topic_ids = [f"Topic {i + 1}" for i in range(len(topic_vectors))]

    sim_df = pd.DataFrame(topic_sim_matrix, index=topic_ids, columns=topic_ids)

    plt.figure(figsize=(10, 8))
    sns.heatmap(sim_df, annot=True, fmt=".2f", cmap="viridis", vmin=-1, vmax=1)
    plt.title(f"Topic Similarity Matrix ({topic})")
    plt.show()


def plot_clustering_dendogram(topic_vectors, topic):
    """
    Plot hierarchical clustering dendrogram for topic vectors.
    
    Args:
        topic_vectors: Array of topic embedding vectors
        topic: Division name for plot title
    """
    dist_matrix = pdist(topic_vectors, metric="cosine")

    Z = linkage(dist_matrix, method="ward")

    topic_ids = [f"Topic {i + 1}" for i in range(len(topic_vectors))]

    plt.figure(figsize=(12, 6))
    dendrogram(Z, labels=topic_ids, leaf_rotation=90)
    plt.axhline(y=0.15, c='red', linestyle='--')
    plt.title(f"Hierarchical Clustering Dendrogram ({topic})")
    plt.xlabel("Topic Index")
    plt.ylabel("Distance (1 - cosine similarity)")
    plt.show()


def plot_inconsistency_diagram(topic_vectors, topic):
    """
    Plot inconsistency diagram for clustering analysis.
    
    Args:
        topic_vectors: Array of topic embedding vectors
        topic: Division name for plot title
    """
    dist_matrix = pdist(topic_vectors, metric="cosine")

    Z = linkage(dist_matrix, method="ward")

    merge_distances = Z[:, 2]
    inc = inconsistent(Z)
    coeff = inc[:, -1]

    plt.figure(figsize=(8, 5))
    plt.scatter(merge_distances, coeff)
    plt.xlabel("Merge distance")
    plt.ylabel("Inconsistency coefficient")
    plt.title(f"Inconsistency vs Merge Distance ({topic})")
    plt.tight_layout()
    plt.show()


def perform_subtopic_clustering(topic_vectors, topic, threshold):
    """
    Perform hierarchical clustering on topic vectors.
    
    Args:
        topic_vectors: Array of topic embedding vectors
        topic: Division name (unused, kept for compatibility)
        threshold: Distance threshold for clustering
        
    Returns:
        Dictionary mapping topic IDs to cluster IDs
    """
    dist_matrix = pdist(topic_vectors, metric="cosine")

    topic_ids = [f"Topic {i + 1}" for i in range(len(topic_vectors))]
    Z = linkage(dist_matrix, method="ward")

    cluster_labels = fcluster(Z, t=threshold, criterion="distance")

    topic_cluster_dict = {}
    for tid, cid in zip(topic_ids, cluster_labels):
        topic_cluster_dict[tid] = cid

    return topic_cluster_dict


def dict_to_multisheet_excel(data, excel_path="clustered_subtopics_by_division.xlsx"):
    """
    Export clustered topics to Excel with one sheet per division.
    
    Args:
        data: Dictionary mapping division names to cluster-keyphrase dictionaries
        excel_path: Output Excel file path
    """
    with pd.ExcelWriter(excel_path, engine="xlsxwriter") as writer:
        
        for division, clusters in data.items():
            
            rows = []
            for cluster_id, words in clusters.items():
                row = {
                    "Cluster": cluster_id,
                }
                for i, w in enumerate(words, start=1):
                    row[f"word_{i}"] = w
                rows.append(row)

            df = pd.DataFrame(rows)
            
            safe_sheet_name = division.replace("/", "_").replace("\\", "_")
            
            df.to_excel(writer, sheet_name=safe_sheet_name[:31], index=False)

    print(f"Saved Excel file to: {excel_path}")


def plot_yearly_cluster_dist(df, cluster_cols, topic, save_path, display_names=None):
    """
    Plot yearly cluster distribution trends.
    
    Args:
        df: DataFrame with year and cluster percentage columns
        cluster_cols: List of cluster column names to plot
        topic: Division name for plot title
        save_path: Path to save the plot
        display_names: Optional dictionary mapping column names to display names
    """
    import textwrap
    
    fig, ax = plt.subplots(figsize=(12,8))

    wrapped_labels = []
    for col in cluster_cols:
        if display_names and col in display_names:
            label = display_names[col]
        else:
            label = col.replace("_", " ")
        wrapped = "\n".join(textwrap.wrap(label, width=50, break_long_words=True))
        wrapped_labels.append(wrapped)
        ax.plot(df["year"], df[col], marker="o", label=wrapped)

    ax.set_xlabel("Year")
    ax.set_ylabel("Percentage")
    ax.set_title(f"Mean Cluster Topic Percentages by Year ({topic})")
    ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.15), ncol=2, fontsize=9, frameon=False)
    ax.grid(True)
    plt.tight_layout()
    plt.savefig(save_path, bbox_inches='tight', dpi=150)
    plt.close()


def main(excel_path, topic_assignments_path, cluster_threshold=0.15):
    """
    Main function to perform subtopic clustering pipeline.
    
    Args:
        excel_path: Path to Excel file with topics by division (lda_topics_by_division.xlsx)
        topic_assignments_path: Path to CSV file with topic assignments
        cluster_threshold: Distance threshold for clustering (default 0.15)
    """
    # Load topic keyphrases from Excel
    print("Loading topic keyphrases from Excel...")
    topic_keyphrase_lists = {}
    topic_name_lists = {}
    wb = load_workbook(excel_path, data_only=True)
    
    for sheet_name in wb.sheetnames:
        if sheet_name == 'master':
            continue
        ws = wb[sheet_name]
        keyphrases = []
        topic_names = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_list = list(row)
            if row_list:
                topic_name = row_list[0] if len(row_list) > 0 else f"Topic {len(topic_names) + 1}"
                topic_names.append(topic_name)
                keyphrases.append(row_list[1:] if len(row_list) > 1 else [])
        
        topic_keyphrase_lists[sheet_name] = keyphrases
        topic_name_lists[sheet_name] = topic_names
    
    # Load sentence transformer model
    print("Loading sentence transformer model...")
    model = SentenceTransformer("sentence-transformers/all-mpnet-base-v2")
    
    # Compute topic embeddings
    print("Computing topic embeddings...")
    topic_embs = {}
    for topic in topic_keyphrase_lists.keys():
        topic_vectors = []
        for keyphrase_list in topic_keyphrase_lists[topic]:
            emb = topic_centroid(keyphrase_list, model)
            topic_vectors.append(emb)
        topic_embs[topic] = topic_vectors
    
    # Perform clustering for all divisions
    print("Performing subtopic clustering...")
    cluster_assignments = {}
    for topic in topic_embs.keys():
        topic_vectors = np.vstack(topic_embs[topic])
        cluster_assignments[topic] = perform_subtopic_clustering(topic_vectors, topic, cluster_threshold)
    
    # Create mapping from topic IDs to actual topic names
    topic_id_to_name = {}
    for division in topic_name_lists:
        for i, topic_name in enumerate(topic_name_lists[division]):
            topic_id = f"Topic {i + 1}"
            topic_id_to_name[(division, topic_id)] = topic_name
    
    # Reorganize keyphrases by cluster
    print("Reorganizing keyphrases by cluster...")
    cluster_topic_keyphrase_lists = {}
    for topic in topic_keyphrase_lists:
        keyphrases = {}
        cluster_to_name = {}
        for tid, cid in cluster_assignments[topic].items():
            if cid not in cluster_to_name:
                cluster_to_name[cid] = []
            cluster_to_name[cid].append(tid)
        
        for cid, topic_ids in cluster_to_name.items():
            actual_topic_names = [topic_id_to_name.get((topic, tid), tid) for tid in sorted(topic_ids)]
            cluster_name = "/".join(actual_topic_names)
            for i, keyphrase_list in enumerate(topic_keyphrase_lists[topic]):
                if f"Topic {i + 1}" in topic_ids:
                    if cluster_name in keyphrases:
                        keyphrases[cluster_name].update(keyphrase_list)
                    else:
                        keyphrases[cluster_name] = set(keyphrase_list)
        
        cluster_topic_keyphrase_lists[topic] = keyphrases
    
    # Export clustered topics to Excel
    print("Exporting clustered topics to Excel...")
    output_excel_path = Path(excel_path).parent / "clustered_subtopics_by_division.xlsx"
    dict_to_multisheet_excel(cluster_topic_keyphrase_lists, str(output_excel_path))
    
    # Reclassify grants with new clusters
    print("Reclassifying grants with new clusters...")
    orig_assignments = pd.read_csv(topic_assignments_path, low_memory=False)
    if len(orig_assignments.columns) > 0 and orig_assignments.columns[0] == 'Unnamed: 0':
        orig_assignments.drop(orig_assignments.columns[0], axis=1, inplace=True)
    
    topic_cols = [col for col in orig_assignments.columns if col.startswith("Topic_")]
    grouped = orig_assignments.groupby(["division", "year"])[topic_cols].mean().reset_index()
    
    for topic in cluster_assignments:
        cluster_ids = set(cluster_assignments[topic].values())
        topic_df_vals = grouped[grouped["division"].str.contains(topic, na=False)].copy()
        
        if len(topic_df_vals) == 0:
            continue
        
        cluster_names = []
        cluster_display_names = {}
        for cid in cluster_ids:
            topic_ids_in_cluster = [tid for tid, c in cluster_assignments[topic].items() if c == cid]
            actual_topic_names = [topic_id_to_name.get((topic, tid), tid) for tid in sorted(topic_ids_in_cluster)]
            cluster_name_display = "/".join(actual_topic_names)
            cols_in_cluster = [f"{tid}_pct".replace(" ", "_") for tid in topic_ids_in_cluster]
            cols_in_cluster = [col for col in cols_in_cluster if col in topic_df_vals.columns]
            
            if cols_in_cluster:
                cluster_column_name = cluster_name_display.replace("/", "_").replace(" ", "_")
                topic_df_vals.loc[:, cluster_column_name] = topic_df_vals[cols_in_cluster].sum(axis=1)
                cluster_names.append(cluster_column_name)
                cluster_display_names[cluster_column_name] = cluster_name_display
        
        cluster_cols = [col for col in cluster_names if col in topic_df_vals.columns]
        
        if cluster_cols:
            os.makedirs(Path(excel_path).parent / "clustered_subtopic_percentages", exist_ok=True)
            plot_path = Path(excel_path).parent / f"clustered_subtopic_percentages/{topic.replace('/', '_')}.png"
            plot_yearly_cluster_dist(topic_df_vals, cluster_cols, topic, save_path=str(plot_path), display_names=cluster_display_names)
    
    print("\nSubtopic clustering complete!")
    return cluster_assignments, cluster_topic_keyphrase_lists

if __name__ == "__main__":
    main(
        excel_path="data/lda_topics_by_division_labeled.xlsx",
        topic_assignments_path="data/topic_assignments.csv",
        cluster_threshold=0.15
    )